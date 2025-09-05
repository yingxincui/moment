#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel报告工具模块
为ETF动量策略分析系统生成Excel格式的分析报告
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

class ExcelReportGenerator:
    """Excel报告生成器"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "ETF动量策略分析报告"
        self._setup_styles()
    
    def _setup_styles(self):
        """设置Excel样式"""
        # 标题样式
        self.title_font = Font(name='微软雅黑', size=16, bold=True, color='1f77b4')
        self.subtitle_font = Font(name='微软雅黑', size=14, bold=True, color='2e7d32')
        self.body_font = Font(name='微软雅黑', size=11, color='424242')
        self.emphasis_font = Font(name='微软雅黑', size=11, bold=True, color='d32f2f')
        
        # 填充样式
        self.title_fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
        self.subtitle_fill = PatternFill(start_color='f3e5f5', end_color='f3e5f5', fill_type='solid')
        self.header_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
        
        # 边框样式
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 对齐样式
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def add_title(self, title, row=1):
        """添加标题"""
        self.ws.merge_cells(f'A{row}:H{row}')
        cell = self.ws[f'A{row}']
        cell.value = title
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_alignment
        return row + 2
    
    def add_subtitle(self, subtitle, row):
        """添加副标题"""
        self.ws.merge_cells(f'A{row}:H{row}')
        cell = self.ws[f'A{row}']
        cell.value = subtitle
        cell.font = self.subtitle_font
        cell.fill = self.subtitle_fill
        cell.alignment = self.left_alignment
        return row + 2
    
    def add_dataframe(self, df, title, row, start_col='A'):
        """添加DataFrame到Excel"""
        # 添加标题
        self.ws[f'{start_col}{row}'] = title
        self.ws[f'{start_col}{row}'].font = self.subtitle_font
        self.ws[f'{start_col}{row}'].fill = self.subtitle_fill
        row += 1
        
        # 添加列标题
        for col_idx, col_name in enumerate(df.columns):
            cell = self.ws[f'{chr(ord(start_col) + col_idx)}{row}']
            cell.value = col_name
            cell.font = self.body_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        row += 1
        
        # 添加数据
        for _, data_row in df.iterrows():
            for col_idx, value in enumerate(data_row):
                cell = self.ws[f'{chr(ord(start_col) + col_idx)}{row}']
                cell.value = value
                cell.font = self.body_font
                cell.border = self.border
                cell.alignment = self.center_alignment
            row += 1
        
        return row + 1
    
    def add_metrics(self, metrics_data, row):
        """添加关键指标"""
        col = 0
        for metric_name, metric_value in metrics_data.items():
            # 指标名称
            name_cell = self.ws[f'{chr(65 + col)}{row}']
            name_cell.value = metric_name
            name_cell.font = self.body_font
            name_cell.fill = self.header_fill
            name_cell.border = self.border
            name_cell.alignment = self.center_alignment
            
            # 指标值
            value_cell = self.ws[f'{chr(65 + col)}{row + 1}']
            value_cell.value = metric_value
            value_cell.font = self.emphasis_font
            value_cell.border = self.border
            value_cell.alignment = self.center_alignment
            
            col += 2
        
        return row + 3
    
    def add_text_content(self, content, row):
        """添加文本内容"""
        self.ws[f'A{row}'] = content
        self.ws[f'A{row}'].font = self.body_font
        self.ws[f'A{row}'].alignment = self.left_alignment
        return row + 1
    
    def generate_momentum_report(self, selected_etfs_result, all_etfs_result, etf_pool, 
                                momentum_period, ma_period, max_positions, 
                                backtest_results=None, trade_history=None):
        """生成动量策略分析报告"""
        current_row = 1
        
        # 添加报告标题
        current_row = self.add_title("ETF动量策略分析报告", current_row)
        
        # 添加生成时间
        current_row = self.add_text_content(f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", current_row)
        current_row += 1
        
        # 添加策略参数
        current_row = self.add_subtitle("策略参数", current_row)
        params_data = {
            '动量周期': f"{momentum_period}天",
            '均线周期': f"{ma_period}天",
            '最大持仓': f"{max_positions}只"
        }
        current_row = self.add_metrics(params_data, current_row)
        
        # 添加推荐持仓
        if selected_etfs_result:
            current_row = self.add_subtitle("推荐持仓", current_row)
            
            # 构建推荐ETF列表
            etf_list = []
            for i, etf_info in enumerate(selected_etfs_result, 1):
                etf_code = etf_info[0]
                etf_name = etf_info[1]
                etf_list.append(f"{i}. {etf_code} - {etf_name}")
            
            for etf_item in etf_list:
                current_row = self.add_text_content(etf_item, current_row)
            
            current_row += 1
        
        # 添加所有ETF排名
        if all_etfs_result:
            current_row = self.add_subtitle("所有ETF动量排名", current_row)
            
            # 创建排名表格
            all_data = []
            for etf in all_etfs_result:
                if len(etf) >= 6:
                    status = "推荐" if etf[5] else "不符合条件"
                    all_data.append({
                        'ETF代码': etf[0],
                        'ETF名称': etf[1],
                        '当前价格': f"{etf[2]:.4f}",
                        '均线价格': f"{etf[3]:.4f}",
                        '动量': f"{etf[4]*100:.2f}%",
                        '状态': status
                    })
            
            if all_data:
                all_df = pd.DataFrame(all_data)
                current_row = self.add_dataframe(all_df, "ETF动量排名表", current_row)
        
        # 添加回测结果
        if backtest_results:
            current_row = self.add_subtitle("回测结果", current_row)
            
            # 关键指标
            metrics_data = {
                '总收益率': f"{backtest_results.get('total_return', 0):.2f}%",
                '年化收益率': f"{backtest_results.get('annual_return', 0):.2f}%",
                '最大回撤': f"{backtest_results.get('max_drawdown', 0):.2f}%",
                '夏普比率': f"{backtest_results.get('sharpe_ratio', 0):.2f}"
            }
            current_row = self.add_metrics(metrics_data, current_row)
            
            # 交易历史
            if trade_history:
                current_row = self.add_subtitle("交易历史", current_row)
                trade_df = pd.DataFrame(trade_history)
                current_row = self.add_dataframe(trade_df, "交易记录", current_row)
        
        # 添加策略说明
        current_row = self.add_subtitle("策略说明", current_row)
        strategy_text = f"""
动量策略逻辑：
1. 动量计算: 计算各ETF在{momentum_period}天内的价格变化百分比
2. 趋势过滤: 使用{ma_period}天移动平均线过滤下跌趋势
3. 持仓选择: 选择动量最强且趋势向上的ETF
4. 动态调整: 定期重新计算并调整持仓
5. 风险控制: 结合价格与均线关系进行风险控制

操作建议：
- 当前持仓：{len(selected_etfs_result) if selected_etfs_result else 0}只ETF
- 建议：可以适当持有第3名ETF作为缓冲
- 调仓时机：关注排名变化，避免频繁交易
        """
        
        for line in strategy_text.strip().split('\n'):
            if line.strip():
                current_row = self.add_text_content(line.strip(), current_row)
        
        # 调整列宽
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
        
        return self.wb
    
    def save_report(self, filename=None):
        """保存报告"""
        if filename is None:
            filename = f"ETF动量策略分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 保存到内存
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output, filename

def generate_excel_report(selected_etfs_result, all_etfs_result, etf_pool, 
                         momentum_period, ma_period, max_positions,
                         backtest_results=None, trade_history=None):
    """生成Excel报告的便捷函数"""
    generator = ExcelReportGenerator()
    wb = generator.generate_momentum_report(
        selected_etfs_result, all_etfs_result, etf_pool,
        momentum_period, ma_period, max_positions,
        backtest_results, trade_history
    )
    return generator.save_report()

def download_excel_report_button(selected_etfs_result, all_etfs_result, etf_pool,
                               momentum_period, ma_period, max_positions,
                               backtest_results=None, trade_history=None,
                               button_text="📊 下载Excel分析报告"):
    """创建下载Excel报告的按钮"""
    try:
        output, filename = generate_excel_report(
            selected_etfs_result, all_etfs_result, etf_pool,
            momentum_period, ma_period, max_positions,
            backtest_results, trade_history
        )
        
        # 创建下载按钮
        st.download_button(
            label=button_text,
            data=output.getvalue(),
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="点击下载Excel格式的分析报告"
        )
        
        return True
    except Exception as e:
        st.error(f"生成Excel报告失败: {e}")
        return False
